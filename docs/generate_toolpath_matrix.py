"""Generate Toolpath Configuration Matrix spreadsheet."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Style definitions ──
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
subheader_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
section_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
section_font = Font(name='Calibri', bold=True, size=11)
normal_font = Font(name='Calibri', size=11)
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_subheader_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = subheader_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_section_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = wrap_align
        cell.border = thin_border

def style_data_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font
        cell.alignment = wrap_align
        cell.border = thin_border

def write_row(ws, row, data):
    for c, val in enumerate(data, 1):
        ws.cell(row=row, column=c, value=val)

# ═══════════════════════════════════════════════════════════════
# SHEET 1: STILE CONFIGURATIONS
# ═══════════════════════════════════════════════════════════════
ws_stile = wb.active
ws_stile.title = 'Stile Configurations'
stile_cols = 8
stile_headers = [
    'Configuration', 'Doors', 'LD Swing', 'RD Swing',
    'Setup 1 (Posted First)', 'Setup 2 (Posted Second)',
    'G-Code Order', 'Notes'
]

row = 1
ws_stile.cell(row=row, column=1, value='STILE TOOLPATH CONFIGURATIONS (3082 / 3086)')
ws_stile.merge_cells(start_row=row, start_column=1, end_row=row, end_column=stile_cols)
ws_stile.cell(row=row, column=1).font = Font(name='Calibri', bold=True, size=14, color='2F5496')
ws_stile.cell(row=row, column=1).alignment = Alignment(horizontal='center')

row = 2
ws_stile.cell(row=row, column=1, value='Each stile setup contains 1 toolpath operation (the rabbet cut). Setup selection determines which rabbet toolpath gets posted.')
ws_stile.merge_cells(start_row=row, start_column=1, end_row=row, end_column=stile_cols)
ws_stile.cell(row=row, column=1).font = Font(name='Calibri', italic=True, size=10, color='666666')

row = 3
ws_stile.cell(row=row, column=1, value='Available setups: Left Rabbet - Out G57, Left Rabbet - Out G59, Left Rabbet - In G57, Left Rabbet - In G59, Right Rabbet - Out G57, Right Rabbet - Out G59, Right Rabbet - In G57, Right Rabbet - In G59')
ws_stile.merge_cells(start_row=row, start_column=1, end_row=row, end_column=stile_cols)
ws_stile.cell(row=row, column=1).font = Font(name='Calibri', italic=True, size=10, color='666666')

# Headers
row = 5
write_row(ws_stile, row, stile_headers)
style_header_row(ws_stile, row, stile_cols)

# ── 1-Door Cases ──
row = 6
write_row(ws_stile, row, ['1-DOOR STILE CASES (single setup posted)', '', '', '', '', '', '', ''])
style_section_row(ws_stile, row, stile_cols)
ws_stile.merge_cells(start_row=row, start_column=1, end_row=row, end_column=stile_cols)

stile_1door = [
    ['Left door, Out-swing',  1, 'O/S', 'N/A', 'Left Rabbet - Out G57',  '—', 'G57 only',  'Exterior rabbet on left side'],
    ['Left door, In-swing',   1, 'I/S', 'N/A', 'Left Rabbet - In G59',   '—', 'G59 only',  'Interior rabbet on left side'],
    ['Right door, Out-swing', 1, 'N/A', 'O/S', 'Right Rabbet - Out G59', '—', 'G59 only',  'Exterior rabbet on right side'],
    ['Right door, In-swing',  1, 'N/A', 'I/S', 'Right Rabbet - In G57',  '—', 'G57 only',  'Interior rabbet on right side'],
]

for data in stile_1door:
    row += 1
    write_row(ws_stile, row, data)
    style_data_row(ws_stile, row, stile_cols)

# ── 2-Door Same-Face Cases ──
row += 1
write_row(ws_stile, row, ['2-DOOR STILE: SAME-FACE RABBETING (both doors swing same direction)', '', '', '', '', '', '', ''])
style_section_row(ws_stile, row, stile_cols)
ws_stile.merge_cells(start_row=row, start_column=1, end_row=row, end_column=stile_cols)

stile_2door_same = [
    ['Both In-swing',  2, 'I/S', 'I/S', 'Right Rabbet - In G57', 'Left Rabbet - In G59',  'G57 → Rotate 180° → G59', 'Interior rabbet both sides.\nFixed order: right G57 first, left G59 second.\nDrilling flags NOT used.'],
    ['Both Out-swing', 2, 'O/S', 'O/S', 'Left Rabbet - Out G57', 'Right Rabbet - Out G59', 'G57 → Rotate 180° → G59', 'Exterior rabbet both sides.\nFixed order: left G57 first, right G59 second.\nDrilling flags NOT used.'],
]

for data in stile_2door_same:
    row += 1
    write_row(ws_stile, row, data)
    style_data_row(ws_stile, row, stile_cols)

# ── 2-Door Opposite-Face Cases ──
row += 1
write_row(ws_stile, row, ['2-DOOR STILE: OPPOSITE-FACE RABBETING (doors swing opposite directions) — Order determined by drilling', '', '', '', '', '', '', ''])
style_section_row(ws_stile, row, stile_cols)
ws_stile.merge_cells(start_row=row, start_column=1, end_row=row, end_column=stile_cols)

row += 1
write_row(ws_stile, row, ['Sub-case: Left In + Right Out (both setups use G59)', '', '', '', '', '', '', ''])
style_subheader_row(ws_stile, row, stile_cols)
ws_stile.merge_cells(start_row=row, start_column=1, end_row=row, end_column=stile_cols)

stile_opp_g59 = [
    ['No drilling (3082)',            2, 'I/S', 'O/S', 'Left Rabbet - In G59',  'Right Rabbet - Out G59', 'G59 → Flip → G59', 'Left first (3082 default)'],
    ['No drilling (3086)',            2, 'I/S', 'O/S', 'Right Rabbet - Out G59', 'Left Rabbet - In G59',  'G59 → Flip → G59', 'Right first (3086 default)'],
    ['Drilling left only',           2, 'I/S', 'O/S', 'Left Rabbet - In G59',  'Right Rabbet - Out G59', 'G59 → Flip → G59', 'Side with drilling goes first'],
    ['Drilling right only',          2, 'I/S', 'O/S', 'Right Rabbet - Out G59', 'Left Rabbet - In G59',  'G59 → Flip → G59', 'Side with drilling goes first'],
    ['Drilling both, left FR=3',     2, 'I/S', 'O/S', 'Left Rabbet - In G59',  'Right Rabbet - Out G59', 'G59 → Flip → G59', 'Left exterior drilling (FR=3) → left first'],
    ['Drilling both, right FR=3',    2, 'I/S', 'O/S', 'Right Rabbet - Out G59', 'Left Rabbet - In G59',  'G59 → Flip → G59', 'Right exterior drilling (FR=3) → right first'],
]

for data in stile_opp_g59:
    row += 1
    write_row(ws_stile, row, data)
    style_data_row(ws_stile, row, stile_cols)

row += 1
write_row(ws_stile, row, ['Sub-case: Left Out + Right In (both setups use G57)', '', '', '', '', '', '', ''])
style_subheader_row(ws_stile, row, stile_cols)
ws_stile.merge_cells(start_row=row, start_column=1, end_row=row, end_column=stile_cols)

stile_opp_g57 = [
    ['No drilling (3082)',            2, 'O/S', 'I/S', 'Left Rabbet - Out G57',  'Right Rabbet - In G57', 'G57 → Flip → G57', 'Left first (3082 default)'],
    ['No drilling (3086)',            2, 'O/S', 'I/S', 'Right Rabbet - In G57',  'Left Rabbet - Out G57', 'G57 → Flip → G57', 'Right first (3086 default)'],
    ['Drilling left only',           2, 'O/S', 'I/S', 'Left Rabbet - Out G57',  'Right Rabbet - In G57', 'G57 → Flip → G57', 'Side with drilling goes first'],
    ['Drilling right only',          2, 'O/S', 'I/S', 'Right Rabbet - In G57',  'Left Rabbet - Out G57', 'G57 → Flip → G57', 'Side with drilling goes first'],
    ['Drilling both, left FR=3',     2, 'O/S', 'I/S', 'Left Rabbet - Out G57',  'Right Rabbet - In G57', 'G57 → Flip → G57', 'Left exterior drilling (FR=3) → left first'],
    ['Drilling both, right FR=3',    2, 'O/S', 'I/S', 'Right Rabbet - In G57',  'Left Rabbet - Out G57', 'G57 → Flip → G57', 'Right exterior drilling (FR=3) → right first'],
]

for data in stile_opp_g57:
    row += 1
    write_row(ws_stile, row, data)
    style_data_row(ws_stile, row, stile_cols)

# Set column widths
stile_widths = [38, 8, 10, 10, 28, 28, 26, 48]
for i, w in enumerate(stile_widths, 1):
    ws_stile.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════
# SHEET 2: DOOR CONFIGURATIONS
# ═══════════════════════════════════════════════════════════════
ws_door = wb.create_sheet('Door Configurations')
door_cols = 9
door_headers = [
    'Configuration', 'Swing', 'Left Features', 'Right Features',
    'Setup Selected', 'G-Code', 'Toolpath: Rabbet', 'Toolpath: Bottom Notch', 'Toolpath: Top Notch'
]

row = 1
ws_door.cell(row=row, column=1, value='DOOR TOOLPATH CONFIGURATIONS')
ws_door.merge_cells(start_row=row, start_column=1, end_row=row, end_column=door_cols)
ws_door.cell(row=row, column=1).font = Font(name='Calibri', bold=True, size=14, color='2F5496')
ws_door.cell(row=row, column=1).alignment = Alignment(horizontal='center')

row = 2
ws_door.cell(row=row, column=1, value='Each door setup has 3 toolpaths: Rabbet, Bottom Notch, Top Notch. Each is independently suppressed/unsuppressed based on feature flags.')
ws_door.merge_cells(start_row=row, start_column=1, end_row=row, end_column=door_cols)
ws_door.cell(row=row, column=1).font = Font(name='Calibri', italic=True, size=10, color='666666')

row = 3
ws_door.cell(row=row, column=1, value='8 setups total: Setup 1-G57 thru Setup 8-G59. Interior vs Exterior determined by rabbeting flag. Notch toolpaths are identical in both interior/exterior setups.')
ws_door.merge_cells(start_row=row, start_column=1, end_row=row, end_column=door_cols)
ws_door.cell(row=row, column=1).font = Font(name='Calibri', italic=True, size=10, color='666666')

row = 5
write_row(ws_door, row, door_headers)
style_header_row(ws_door, row, door_cols)

# ── Door Setup Map ──
row = 6
write_row(ws_door, row, ['SETUP MAP (which setup is selected based on side + swing + rabbet type)', '', '', '', '', '', '', '', ''])
style_section_row(ws_door, row, door_cols)
ws_door.merge_cells(start_row=row, start_column=1, end_row=row, end_column=door_cols)

door_setup_map = [
    ['Setup 1 - G57', '', 'Left side', '', 'Left Interior Rabbet', 'G57', 'Left Interior Rabbet', 'Left Bottom Notch', 'Left Top Notch'],
    ['Setup 2 - G59', '', 'Left side', '', 'Left Interior Rabbet', 'G59', 'Left Interior Rabbet', 'Left Bottom Notch', 'Left Top Notch'],
    ['Setup 3 - G57', '', 'Left side', '', 'Left Exterior Rabbet', 'G57', 'Left Exterior Rabbet', 'Left Bottom Notch', 'Left Top Notch'],
    ['Setup 4 - G59', '', 'Left side', '', 'Left Exterior Rabbet', 'G59', 'Left Exterior Rabbet', 'Left Bottom Notch', 'Left Top Notch'],
    ['Setup 5 - G57', '', '', 'Right side', 'Right Interior Rabbet', 'G57', 'Right Interior Rabbet', 'Right Bottom Notch', 'Right Top Notch'],
    ['Setup 6 - G59', '', '', 'Right side', 'Right Interior Rabbet', 'G59', 'Right Interior Rabbet', 'Right Bottom Notch', 'Right Top Notch'],
    ['Setup 7 - G57', '', '', 'Right side', 'Right Exterior Rabbet', 'G57', 'Right Exterior Rabbet', 'Right Bottom Notch', 'Right Top Notch'],
    ['Setup 8 - G59', '', '', 'Right side', 'Right Exterior Rabbet', 'G59', 'Right Exterior Rabbet', 'Right Bottom Notch', 'Right Top Notch'],
]

for data in door_setup_map:
    row += 1
    write_row(ws_door, row, data)
    style_data_row(ws_door, row, door_cols)

# ── Door Selection Logic ──
row += 1
write_row(ws_door, row, ['SETUP SELECTION LOGIC (which setups get posted and in what order)', '', '', '', '', '', '', '', ''])
style_section_row(ws_door, row, door_cols)
ws_door.merge_cells(start_row=row, start_column=1, end_row=row, end_column=door_cols)

row += 1
write_row(ws_door, row, ['OUT-SWING DOOR', '', '', '', '', '', '', '', ''])
style_subheader_row(ws_door, row, door_cols)
ws_door.merge_cells(start_row=row, start_column=1, end_row=row, end_column=door_cols)

door_out = [
    ['Left only, exterior rabbet',   'O/S', 'Ext rabbet + notches',  '—',                    'Setup 4 - G59', 'G59 only',          'Active if left_exterior_rabbeting',  'Active if bottom_left_notching', 'Active if top_left_notching'],
    ['Left only, interior rabbet',   'O/S', 'Int rabbet + notches',  '—',                    'Setup 2 - G59', 'G59 only',          'Active if left_interior_rabbeting',  'Active if bottom_left_notching', 'Active if top_left_notching'],
    ['Left only, notches only',      'O/S', 'Notches only',          '—',                    'Setup 2 - G59', 'G59 only',          'Suppressed (no rabbet)',             'Active if bottom_left_notching', 'Active if top_left_notching'],
    ['Right only, exterior rabbet',  'O/S', '—',                     'Ext rabbet + notches', 'Setup 7 - G57', 'G57 only',          'Active if right_exterior_rabbeting', 'Active if bottom_right_notching', 'Active if top_right_notching'],
    ['Right only, interior rabbet',  'O/S', '—',                     'Int rabbet + notches', 'Setup 5 - G57', 'G57 only',          'Active if right_interior_rabbeting', 'Active if bottom_right_notching', 'Active if top_right_notching'],
    ['Both sides',                   'O/S', 'Features present',      'Features present',     'Left G59, then Right G57', 'G59 → G57', 'Per side rabbet flag',            'Per side notch flag',             'Per side notch flag'],
]

for data in door_out:
    row += 1
    write_row(ws_door, row, data)
    style_data_row(ws_door, row, door_cols)

row += 1
write_row(ws_door, row, ['IN-SWING DOOR', '', '', '', '', '', '', '', ''])
style_subheader_row(ws_door, row, door_cols)
ws_door.merge_cells(start_row=row, start_column=1, end_row=row, end_column=door_cols)

door_in = [
    ['Left only, exterior rabbet',   'I/S', 'Ext rabbet + notches',  '—',                    'Setup 3 - G57', 'G57 only',          'Active if left_exterior_rabbeting',  'Active if bottom_left_notching', 'Active if top_left_notching'],
    ['Left only, interior rabbet',   'I/S', 'Int rabbet + notches',  '—',                    'Setup 1 - G57', 'G57 only',          'Active if left_interior_rabbeting',  'Active if bottom_left_notching', 'Active if top_left_notching'],
    ['Left only, notches only',      'I/S', 'Notches only',          '—',                    'Setup 1 - G57', 'G57 only',          'Suppressed (no rabbet)',             'Active if bottom_left_notching', 'Active if top_left_notching'],
    ['Right only, exterior rabbet',  'I/S', '—',                     'Ext rabbet + notches', 'Setup 8 - G59', 'G59 only',          'Active if right_exterior_rabbeting', 'Active if bottom_right_notching', 'Active if top_right_notching'],
    ['Right only, interior rabbet',  'I/S', '—',                     'Int rabbet + notches', 'Setup 6 - G59', 'G59 only',          'Active if right_interior_rabbeting', 'Active if bottom_right_notching', 'Active if top_right_notching'],
    ['Both sides',                   'I/S', 'Features present',      'Features present',     'Left G57, then Right G59', 'G57 → G59', 'Per side rabbet flag',            'Per side notch flag',             'Per side notch flag'],
]

for data in door_in:
    row += 1
    write_row(ws_door, row, data)
    style_data_row(ws_door, row, door_cols)

# ── Suppression Rules ──
row += 2
write_row(ws_door, row, ['TOOLPATH SUPPRESSION RULES (applied after toolpath regeneration, before post-processing)', '', '', '', '', '', '', '', ''])
style_section_row(ws_door, row, door_cols)
ws_door.merge_cells(start_row=row, start_column=1, end_row=row, end_column=door_cols)

supp_headers = ['Toolpath Operation', 'Active When', '', '', '', '', '', '', '']
row += 1
write_row(ws_door, row, supp_headers)
style_subheader_row(ws_door, row, door_cols)

supp_data = [
    ['Left Interior Rabbet',  'left_interior_rabbeting = 1', '', '', '', '', '', '', ''],
    ['Left Exterior Rabbet',  'left_exterior_rabbeting = 1', '', '', '', '', '', '', ''],
    ['Right Interior Rabbet', 'right_interior_rabbeting = 1', '', '', '', '', '', '', ''],
    ['Right Exterior Rabbet', 'right_exterior_rabbeting = 1', '', '', '', '', '', '', ''],
    ['Left Bottom Notch',     'bottom_left_notching ≠ 0', '', '', '', '', '', '', ''],
    ['Left Top Notch',        'top_left_notching ≠ 0', '', '', '', '', '', '', ''],
    ['Right Bottom Notch',    'bottom_right_notching ≠ 0', '', '', '', '', '', '', ''],
    ['Right Top Notch',       'top_right_notching ≠ 0', '', '', '', '', '', '', ''],
]

for data in supp_data:
    row += 1
    write_row(ws_door, row, data)
    style_data_row(ws_door, row, door_cols)

door_widths = [35, 8, 22, 22, 28, 16, 30, 30, 28]
for i, w in enumerate(door_widths, 1):
    ws_door.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════
# SHEET 3: PANEL CONFIGURATIONS
# ═══════════════════════════════════════════════════════════════
ws_panel = wb.create_sheet('Panel Configurations')
panel_cols = 7
panel_headers = [
    'Configuration', 'Orientation', 'Setup Selected', 'G-Code',
    'Toolpath: Notches', 'Toolpath: Cutouts', 'Notes'
]

row = 1
ws_panel.cell(row=row, column=1, value='PANEL TOOLPATH CONFIGURATIONS')
ws_panel.merge_cells(start_row=row, start_column=1, end_row=row, end_column=panel_cols)
ws_panel.cell(row=row, column=1).font = Font(name='Calibri', bold=True, size=14, color='2F5496')
ws_panel.cell(row=row, column=1).alignment = Alignment(horizontal='center')

row = 2
ws_panel.cell(row=row, column=1, value='Each panel setup has up to 6 toolpaths: 4 notches (Front/Back × Top/Bottom) + 2 cutouts (A, B). All at G58. Setup chosen by orientation.')
ws_panel.merge_cells(start_row=row, start_column=1, end_row=row, end_column=panel_cols)
ws_panel.cell(row=row, column=1).font = Font(name='Calibri', italic=True, size=10, color='666666')

row = 4
write_row(ws_panel, row, panel_headers)
style_header_row(ws_panel, row, panel_cols)

# ── Panel Setup Selection ──
row = 5
write_row(ws_panel, row, ['SETUP SELECTION (based on panel orientation)', '', '', '', '', '', ''])
style_section_row(ws_panel, row, panel_cols)
ws_panel.merge_cells(start_row=row, start_column=1, end_row=row, end_column=panel_cols)

panel_setups = [
    ['Portrait (height ≥ width)',  'Height ≥ Width', 'Setup Back Bottom Corner G58', 'G58', 'See suppression rules', 'See suppression rules', 'Single setup posted'],
    ['Landscape (width > height)', 'Width > Height', 'Setup Back Top Corner G58',    'G58', 'See suppression rules', 'See suppression rules', 'Single setup posted'],
    ['No features',                'Any',            '— (skipped)',                   '—',   '—',                    '—',                    'No G-code generated'],
]

for data in panel_setups:
    row += 1
    write_row(ws_panel, row, data)
    style_data_row(ws_panel, row, panel_cols)

# ── Panel Suppression ──
row += 2
write_row(ws_panel, row, ['TOOLPATH SUPPRESSION RULES (applied in BOTH panel setups)', '', '', '', '', '', ''])
style_section_row(ws_panel, row, panel_cols)
ws_panel.merge_cells(start_row=row, start_column=1, end_row=row, end_column=panel_cols)

p_supp_headers = ['Toolpath Operation', 'Active When', '', '', '', '', '']
row += 1
write_row(ws_panel, row, p_supp_headers)
style_subheader_row(ws_panel, row, panel_cols)

panel_supp = [
    ['Front Bottom Notch', 'notching_front_edge_bottom ≠ 0', '', '', '', '', ''],
    ['Front Top Notch',    'notching_front_edge_top ≠ 0',    '', '', '', '', ''],
    ['Back Bottom Notch',  'notching_back_edge_bottom ≠ 0',  '', '', '', '', ''],
    ['Back Top Notch',     'notching_back_edge_top ≠ 0',     '', '', '', '', ''],
    ['Cutout A',           'cutout_A_width > 0 AND cutout_A_height > 0', '', '', '', '', ''],
    ['Cutout B',           'cutout_B_width > 0 AND cutout_B_height > 0', '', '', '', '', ''],
]

for data in panel_supp:
    row += 1
    write_row(ws_panel, row, data)
    style_data_row(ws_panel, row, panel_cols)

panel_widths = [30, 20, 32, 10, 40, 40, 30]
for i, w in enumerate(panel_widths, 1):
    ws_panel.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════
# SHEET 4: SUMMARY / LEGEND
# ═══════════════════════════════════════════════════════════════
ws_legend = wb.create_sheet('Legend & Summary')
leg_cols = 4

row = 1
ws_legend.cell(row=row, column=1, value='LEGEND & SUMMARY')
ws_legend.merge_cells(start_row=row, start_column=1, end_row=row, end_column=leg_cols)
ws_legend.cell(row=row, column=1).font = Font(name='Calibri', bold=True, size=14, color='2F5496')

row = 3
write_row(ws_legend, row, ['Term', 'Meaning', '', ''])
style_header_row(ws_legend, row, leg_cols)

legend = [
    ['G57', 'Work coordinate offset — typically "first side" / "A side" on Anderson Stratos'],
    ['G58', 'Work coordinate offset — used for panels (single-side machining)'],
    ['G59', 'Work coordinate offset — typically "second side" / "B side" on Anderson Stratos'],
    ['I/S (In-swing)', 'Door swings inward — rabbeting on interior face of stile/door'],
    ['O/S (Out-swing)', 'Door swings outward — rabbeting on exterior face of stile/door'],
    ['Rotate 180°', 'Part is rotated on the CNC between setups (same face, opposite end)'],
    ['Flip', 'Part is flipped over on the CNC between setups (opposite face)'],
    ['FlipRotate=3', 'Exterior drilling face — processed FIRST at Gannomat drill machine'],
    ['FlipRotate=4', 'Interior drilling face — processed SECOND at Gannomat drill machine'],
    ['Same-face rabbeting', 'Both doors swing same direction → rabbeting on same face (both interior or both exterior)'],
    ['Opposite-face rabbeting', 'Doors swing opposite directions → rabbeting on different faces (one interior, one exterior)'],
    ['Suppression', 'Fusion CAM operation is set to isSuppressed=True — toolpath exists but is excluded from G-code'],
]

for term, meaning in legend:
    row += 1
    write_row(ws_legend, row, [term, meaning, '', ''])
    style_data_row(ws_legend, row, leg_cols)

row += 2
write_row(ws_legend, row, ['Component', 'G-Codes Used', 'Setups in Model', 'Suppression Logic'])
style_header_row(ws_legend, row, leg_cols)

summary = [
    ['Stile (3082/3086)', 'G57, G59', '8 setups × 1 toolpath each', 'None — setup selection determines which toolpath is posted'],
    ['Door', 'G57, G59', '8 setups × 3 toolpaths each', 'Per-toolpath suppression based on rabbeting + notching flags'],
    ['Panel', 'G58', '2 setups × 6 toolpaths each', 'Per-toolpath suppression based on notching + cutout flags'],
]

for data in summary:
    row += 1
    write_row(ws_legend, row, data)
    style_data_row(ws_legend, row, leg_cols)

legend_widths = [28, 70, 28, 55]
for i, w in enumerate(legend_widths, 1):
    ws_legend.column_dimensions[get_column_letter(i)].width = w


# ── Save ──
output_path = r'c:\Users\james.derrod\VM Fusion Extension\docs\Toolpath_Configuration_Matrix.xlsx'
wb.save(output_path)
print(f'Saved: {output_path}')
