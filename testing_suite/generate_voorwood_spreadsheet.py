"""
Generate Voorwood Testing Planning spreadsheet for STILE cutlist CSV mapping.
Creates an Excel file with input parameters (changeable) and output formulas.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Stile Cutlist Mapping"

# ── Styles ──────────────────────────────────────────────────────────────────
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
section_font = Font(bold=True, size=11, color="FFFFFF")
section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # light yellow
output_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")  # light blue
formula_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # light green
note_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
bold_font = Font(bold=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header(ws, row, cols=5):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

def style_section(ws, row, cols=5):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

def style_row(ws, row, fill, cols=5):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.border = thin_border

# ── Column widths ───────────────────────────────────────────────────────────
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 45
ws.column_dimensions['E'].width = 55

# ============================================================================
# TITLE
# ============================================================================
r = 1
ws.cell(row=r, column=1, value="Voorwood Testing Planning (STILE)")
ws.cell(row=r, column=1).font = Font(bold=True, size=14, color="2F5496")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

r = 2
ws.cell(row=r, column=1, value="Change yellow INPUT cells to see how CSV outputs update via formulas")
ws.cell(row=r, column=1).font = Font(italic=True, size=10, color="666666")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

# ============================================================================
# SECTION 1: INPUT PARAMETERS (F360 Model Variables)
# ============================================================================
r = 4
ws.cell(row=r, column=1, value="F360 MODEL INPUT PARAMETERS")
ws.cell(row=r, column=2, value="VALUE")
ws.cell(row=r, column=3, value="DATA TYPE")
ws.cell(row=r, column=4, value="DESCRIPTION")
ws.cell(row=r, column=5, value="EXAMPLE / NOTES")
style_header(ws, r)

# Input parameters — these are the yellow editable cells
inputs = [
    # (label, default_value, datatype, description, notes)
    ("component_height", 111.9375, "float (in)", "Height of the stile component", "Maps to Length in CSV"),
    ("component_width", 6.0, "float (in)", "Width of the stile component", "Maps to Width in CSV"),
    ("component_thickness", 0.75, "float (in)", "Thickness of the stile component", "Maps to Thickness in CSV"),
    ("ID", "1-S1-IBUS12345", "string", "Full component ID: {qty}-{comp_id}-{SO}", "Format: qty-compID-orderSO"),
    ("Series", "3082G.67P", "string", "Series identifier", "e.g. 3082G.67P, 3086G, etc."),
    ("left_side_door", 1, "bool (0/1)", "Is there a door on the left side?", "1=yes, 0=no"),
    ("left_door_hinging_right", 0, "bool (0/1)", "Left door hinges on right (=this stile is hinge stile)", "1=RH, 0=LH"),
    ("left_door_swinging_out", 0, "bool (0/1)", "Left door swings outward", "1=O/S, 0=I/S"),
    ("left_door_floor_clearance", 4.5, "float (in)", "Floor clearance for left door", "Maps to FC SIDE 2"),
    ("right_side_door", 1, "bool (0/1)", "Is there a door on the right side?", "1=yes, 0=no"),
    ("right_door_hinging_right", 1, "bool (0/1)", "Right door hinges on right (=this stile is keep stile)", "1=RH, 0=LH"),
    ("right_door_swinging_out", 0, "bool (0/1)", "Right door swings outward", "1=O/S, 0=I/S"),
    ("right_door_floor_clearance", 4.5, "float (in)", "Floor clearance for right door", "Maps to FC SIDE 1"),
    ("right_interior_rabbeting", 1, "bool (0/1)", "Right side has interior (bottom face) rabbeting", "Used for GappingSide1 / GappingTopSide1"),
    ("right_exterior_rabbeting", 0, "bool (0/1)", "Right side has exterior (top face) rabbeting", "Used for GappingSide1 / GappingTopSide1"),
    ("left_interior_rabbeting", 0, "bool (0/1)", "Left side has interior (bottom face) rabbeting", "Used for GappingSide2 / GappingTopSide2"),
    ("left_exterior_rabbeting", 1, "bool (0/1)", "Left side has exterior (top face) rabbeting", "Used for GappingSide2 / GappingTopSide2"),
    ("right_rabbeting_top", 2.0, "float (in)", "Right side rabbeting start (from top of stile)", "Maps to GappingStartSide1"),
    ("right_rabbeting_length", 75.0, "float (in)", "Right side rabbeting length", "Maps to GappingLengthSide1"),
    ("left_rabbeting_top", 2.0, "float (in)", "Left side rabbeting start (from top of stile)", "Maps to GappingStartSide2"),
    ("left_rabbeting_length", 75.0, "float (in)", "Left side rabbeting length", "Maps to GappingLengthSide2"),
]

# Write inputs starting at row 5
input_start = 5
for i, (label, val, dtype, desc, notes) in enumerate(inputs):
    r = input_start + i
    ws.cell(row=r, column=1, value=label).font = bold_font
    ws.cell(row=r, column=2, value=val)
    ws.cell(row=r, column=3, value=dtype)
    ws.cell(row=r, column=4, value=desc)
    ws.cell(row=r, column=5, value=notes)
    style_row(ws, r, input_fill)
    # Make value cell stand out more
    ws.cell(row=r, column=2).font = Font(bold=True, size=11)

input_end = input_start + len(inputs) - 1

# Build a dict of input label -> cell reference for formulas
input_cells = {}
for i, (label, _, _, _, _) in enumerate(inputs):
    input_cells[label] = f"B{input_start + i}"

# ============================================================================
# SECTION 2: CSV OUTPUT VARIABLES (with formulas referencing inputs)
# ============================================================================
r = input_end + 2
output_section_start = r
ws.cell(row=r, column=1, value="VOORWOOD CSV OUTPUT (gCutlist)")
ws.cell(row=r, column=2, value="FORMULA VALUE")
ws.cell(row=r, column=3, value="MAPPING TYPE")
ws.cell(row=r, column=4, value="FORMULA / LOGIC")
ws.cell(row=r, column=5, value="CODE REFERENCE (parameter_exporter.py)")
style_header(ws, r)

# Cell references
c_height = input_cells["component_height"]
c_width = input_cells["component_width"]
c_thickness = input_cells["component_thickness"]
c_id = input_cells["ID"]
c_series = input_cells["Series"]
c_left_door = input_cells["left_side_door"]
c_left_hinging = input_cells["left_door_hinging_right"]
c_left_swinging = input_cells["left_door_swinging_out"]
c_left_fc = input_cells["left_door_floor_clearance"]
c_right_door = input_cells["right_side_door"]
c_right_hinging = input_cells["right_door_hinging_right"]
c_right_swinging = input_cells["right_door_swinging_out"]
c_right_fc = input_cells["right_door_floor_clearance"]
c_right_int_rab = input_cells["right_interior_rabbeting"]
c_right_ext_rab = input_cells["right_exterior_rabbeting"]
c_left_int_rab = input_cells["left_interior_rabbeting"]
c_left_ext_rab = input_cells["left_exterior_rabbeting"]
c_right_rab_top = input_cells["right_rabbeting_top"]
c_right_rab_length = input_cells["right_rabbeting_length"]
c_left_rab_top = input_cells["left_rabbeting_top"]
c_left_rab_length = input_cells["left_rabbeting_length"]

# Output rows: (csv_variable, formula, mapping_type, logic_description, code_ref)
outputs = [
    (
        "gCutlist.Item[0].Qty",
        "1",  # constant
        "Constant",
        "Always 1",
        "Line 1033: ('Qty', '1')"
    ),
    (
        "gCutlist.Item[0].Width",
        f"={c_width}",
        "Direct",
        "component_width",
        "Line 903: width_param → component_width"
    ),
    (
        "gCutlist.Item[0].Length",
        f"={c_height}",
        "Direct",
        "component_height (stile height = CSV Length)",
        "Line 906: height_param → component_height"
    ),
    (
        "gCutlist.Item[0].ID",
        f'=MID({c_id},FIND("-",{c_id})+1,FIND("-",{c_id},FIND("-",{c_id})+1)-FIND("-",{c_id})-1)',
        "Extracted",
        'Extract component ID from full ID (e.g. "S1" from "1-S1-IBUS12345")',
        "Line 1036: component_id"
    ),
    (
        "gCutlist.Item[0].Series",
        f"={c_series}",
        "Direct",
        "Series ID passed through",
        "Line 1037: series_id"
    ),
    (
        "gCutlist.Item[0].RoutingInformation",
        # Build: IF left_side_door, "{series} {LH|RH} {IS|OS}"; IF right_side_door, same; join with " / "
        f'=IF(AND({c_left_door}=1,{c_right_door}=1),'
        f'{c_series}&" "&IF({c_left_hinging}=1,"RH","LH")&" "&IF({c_left_swinging}=1,"OS","IS")&" / "&{c_series}&" "&IF({c_right_hinging}=1,"RH","LH")&" "&IF({c_right_swinging}=1,"OS","IS"),'
        f'IF({c_left_door}=1,'
        f'{c_series}&" "&IF({c_left_hinging}=1,"RH","LH")&" "&IF({c_left_swinging}=1,"OS","IS"),'
        f'IF({c_right_door}=1,'
        f'{c_series}&" "&IF({c_right_hinging}=1,"RH","LH")&" "&IF({c_right_swinging}=1,"OS","IS"),'
        f'"")))',
        "Derived",
        '{series} {RH|LH} {OS|IS} per side, joined with " / " if both sides have doors',
        "Lines 939-954: routing_parts built per side"
    ),
    (
        "gCutlist.Item[0].SO",
        f'=MID({c_id},FIND("-",{c_id},FIND("-",{c_id})+1)+1,LEN({c_id}))',
        "Extracted",
        'Extract order/SO from full ID (e.g. "IBUS12345" from "1-S1-IBUS12345")',
        "Line 1039: order_id"
    ),
    (
        "gCutlist.Item[0].FC SIDE 1",
        f'=IF({c_right_door}=1,{c_right_fc},"")',
        "Direct",
        "right_door_floor_clearance (Side1 = right edge)",
        "Lines 914-929: FC from whichever side has a door"
    ),
    (
        "gCutlist.Item[0].FC SIDE 2",
        f'=IF({c_left_door}=1,{c_left_fc},"")',
        "Direct",
        "left_door_floor_clearance (Side2 = left edge)",
        "Lines 914-929: FC from whichever side has a door"
    ),
    (
        "gCutlist.Item[0].GappingSide1",
        f'=IF({c_right_door}=1,1,0)',
        "Derived",
        "1 if right side has a door (= has rabbeting), else 0. Side1 = right edge.",
        "Line 966: gapping_side1 = 1 if right_side_door"
    ),
    (
        "gCutlist.Item[0].GappingSide2",
        f'=IF({c_left_door}=1,1,0)',
        "Derived",
        "1 if left side has a door (= has rabbeting), else 0. Side2 = left edge.",
        "Line 967: gapping_side2 = 1 if left_side_door"
    ),
    (
        "gCutlist.Item[0].GappingSide3",
        "0",
        "Constant",
        "Always 0 (top edge — no rabbeting)",
        "Line 974: gapping_side3 = 0"
    ),
    (
        "gCutlist.Item[0].GappingSide4",
        "0",
        "Constant",
        "Always 0 (bottom edge — no rabbeting)",
        "Line 975: gapping_side4 = 0"
    ),
    (
        "gCutlist.Item[0].GappingStartSide1",
        f'=IF({c_right_door}=1,{c_right_rab_top},0)',
        "Conditional",
        "right_rabbeting_top (start of rabbeting from top) if right side has door",
        "Line 986: right_rabbeting_top if gapping_side1"
    ),
    (
        "gCutlist.Item[0].GappingStartSide2",
        f'=IF({c_left_door}=1,{c_left_rab_top},0)',
        "Conditional",
        "left_rabbeting_top (start of rabbeting from top) if left side has door",
        "Line 989: left_rabbeting_top if gapping_side2"
    ),
    (
        "gCutlist.Item[0].GappingLengthSide1",
        f'=IF({c_right_door}=1,{c_right_rab_length},0)',
        "Conditional",
        "right_rabbeting_length if right side has door",
        "Line 987: right_rabbeting_length if gapping_side1"
    ),
    (
        "gCutlist.Item[0].GappingLengthSide2",
        f'=IF({c_left_door}=1,{c_left_rab_length},0)',
        "Conditional",
        "left_rabbeting_length if left side has door",
        "Line 990: left_rabbeting_length if gapping_side2"
    ),
    (
        "gCutlist.Item[0].GappingStartSide3",
        "0",
        "Constant",
        "N/A (GappingSide3 is always 0)",
        "Line 998: gapping_start_side3 = 0"
    ),
    (
        "gCutlist.Item[0].GappingStartSide4",
        "0",
        "Constant",
        "N/A (GappingSide4 is always 0)",
        "Line 999: gapping_start_side4 = 0"
    ),
    (
        "gCutlist.Item[0].GappingLengthSide3",
        "0",
        "Constant",
        "N/A (GappingSide3 is always 0)",
        "Line 1000: gapping_length_side3 = 0"
    ),
    (
        "gCutlist.Item[0].GappingLengthSide4",
        "0",
        "Constant",
        "N/A (GappingSide4 is always 0)",
        "Line 1001: gapping_length_side4 = 0"
    ),
    (
        "gCutlist.Item[0].GappingTopSide1",
        f'=IF(AND({c_right_door}=1,{c_right_ext_rab}=1),1,0)',
        "Conditional",
        "1 if right side has exterior rabbeting (top face), else 0",
        "Lines 1006-1009: right_exterior_rabbeting if gapping_side1"
    ),
    (
        "gCutlist.Item[0].GappingTopSide2",
        f'=IF(AND({c_left_door}=1,{c_left_ext_rab}=1),1,0)',
        "Conditional",
        "1 if left side has exterior rabbeting (top face), else 0",
        "Lines 1011-1014: left_exterior_rabbeting if gapping_side2"
    ),
    (
        "gCutlist.Item[0].GappingTopSide3",
        "0",
        "Constant",
        "N/A (GappingSide3 is always 0)",
        "Line 1025: gapping_top_side3 = 0"
    ),
    (
        "gCutlist.Item[0].GappingTopSide4",
        "0",
        "Constant",
        "N/A (GappingSide4 is always 0)",
        "Line 1026: gapping_top_side4 = 0"
    ),
    (
        "gCutlist.Item[0].EdgeFinishingSide1",
        f'=IF({c_right_door}=1,0,1)',
        "Derived",
        "Inverse of GappingSide1: 0 if rabbeting on right, else 1",
        "Line 978: edge_finishing_side1 = 0 if gapping_side1 else 1"
    ),
    (
        "gCutlist.Item[0].EdgeFinishingSide2",
        f'=IF({c_left_door}=1,0,1)',
        "Derived",
        "Inverse of GappingSide2: 0 if rabbeting on left, else 1",
        "Line 979: edge_finishing_side2 = 0 if gapping_side2 else 1"
    ),
    (
        "gCutlist.Item[0].EdgeFinishingSide3",
        "1",
        "Constant",
        "Always 1 (top edge always edge-finished)",
        "Line 980: edge_finishing_side3 = 1"
    ),
    (
        "gCutlist.Item[0].EdgeFinishingSide4",
        "1",
        "Constant",
        "Always 1 (bottom edge always edge-finished)",
        "Line 981: edge_finishing_side4 = 1"
    ),
    (
        "gCutlist.Item[0].Thickness",
        f"={c_thickness}",
        "Direct",
        "component_thickness",
        "Line 1029: component_thickness"
    ),
    (
        "gCutlist.Item[0].Flip",
        # Extract comp_id letter from full ID, D=Rotate(1), S=Flip(2) based on drilling flags
        # For stiles: if opposite drilling (left_int XOR right_int, etc) → 2 (Flip), else 1 (Rotate)
        f'=IF(OR(AND({c_left_int_rab}=1,{c_right_ext_rab}=1),AND({c_left_ext_rab}=1,{c_right_int_rab}=1)),2,1)',
        "Derived",
        "Stile: 2 (Flip) if opposite drilling sides (left_int+right_ext or left_ext+right_int), else 1 (Rotate)",
        "Lines 735-748: opposite drilling → FlipRotate=2"
    ),
    (
        "gCutlist.Item[0].Location",
        "BLA",
        "Constant",
        "Always BLA",
        "Line 1063: ('Location', 'BLA')"
    ),
    (
        "gCutlist.Item[0].ProgramName",
        # Same as RoutingInformation for stiles (series + hinge/swing per side)
        f'=IF(AND({c_left_door}=1,{c_right_door}=1),'
        f'{c_series}&" "&IF({c_left_hinging}=1,"RH","LH")&" "&IF({c_left_swinging}=1,"O/S","I/S")&" "&IF({c_left_hinging}=1,"HINGE","KEEP")&", "&IF({c_right_hinging}=1,"RH","LH")&" "&IF({c_right_swinging}=1,"O/S","I/S")&" "&IF({c_right_hinging}=0,"HINGE","KEEP"),'
        f'IF({c_left_door}=1,'
        f'{c_series}&" "&IF({c_left_hinging}=1,"RH","LH")&" "&IF({c_left_swinging}=1,"O/S","I/S")&" "&IF({c_left_hinging}=1,"HINGE","KEEP"),'
        f'IF({c_right_door}=1,'
        f'{c_series}&" "&IF({c_right_hinging}=1,"RH","LH")&" "&IF({c_right_swinging}=1,"O/S","I/S")&" "&IF({c_right_hinging}=0,"HINGE","KEEP"),'
        f'"")))',
        "Derived",
        '{series} {RH|LH} {O/S|I/S} {HINGE|KEEP} per side (HINGE if door hinges onto this stile)',
        "Lines 686-717: DrillName = combined stile drill segments"
    ),
]

# Write outputs
output_start = output_section_start + 1
for i, (csv_var, formula, map_type, logic, code_ref) in enumerate(outputs):
    r = output_start + i
    ws.cell(row=r, column=1, value=csv_var).font = bold_font
    ws.cell(row=r, column=2, value=formula if not formula.startswith("=") else None)
    if formula.startswith("="):
        ws.cell(row=r, column=2).value = formula  # Set as formula
    ws.cell(row=r, column=3, value=map_type)
    ws.cell(row=r, column=4, value=logic)
    ws.cell(row=r, column=5, value=code_ref)
    
    # Color based on mapping type
    if map_type == "Constant":
        style_row(ws, r, note_fill)
    elif map_type == "Direct":
        style_row(ws, r, output_fill)
    elif map_type in ("Derived", "Conditional", "Extracted"):
        style_row(ws, r, formula_fill)
    else:
        style_row(ws, r, output_fill)

output_end = output_start + len(outputs) - 1

# ============================================================================
# SECTION 3: SIDE MAPPING REFERENCE
# ============================================================================
r = output_end + 2
ws.cell(row=r, column=1, value="SIDE MAPPING REFERENCE")
style_section(ws, r)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

ref_data = [
    ("Side1", "Right edge (looking from outside / room side)", "right_* parameters"),
    ("Side2", "Left edge (looking from outside / room side)", "left_* parameters"),
    ("Side3", "Top edge", "Always 0 gapping, always 1 edge finishing"),
    ("Side4", "Bottom edge", "Always 0 gapping, always 1 edge finishing"),
]

r += 1
ws.cell(row=r, column=1, value="Side").font = bold_font
ws.cell(row=r, column=2, value="Physical Location").font = bold_font
ws.cell(row=r, column=3, value="Source Parameters").font = bold_font
style_row(ws, r, note_fill)

for side, loc, src in ref_data:
    r += 1
    ws.cell(row=r, column=1, value=side)
    ws.cell(row=r, column=2, value=loc)
    ws.cell(row=r, column=3, value=src)
    style_row(ws, r, note_fill)

# ============================================================================
# SECTION 4: ROUTING INFO / PROGRAM NAME LOGIC REFERENCE
# ============================================================================
r += 2
ws.cell(row=r, column=1, value="ROUTING / PROGRAM NAME LOGIC")
style_section(ws, r)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

r += 1
ws.cell(row=r, column=1, value="Field").font = bold_font
ws.cell(row=r, column=2, value="Pattern").font = bold_font
ws.cell(row=r, column=3, value="Example").font = bold_font
ws.cell(row=r, column=4, value="Notes").font = bold_font
style_row(ws, r, note_fill)

routing_ref = [
    ("RoutingInformation (1 door)", "{series} {RH|LH} {OS|IS}", "3082G.67P LH IS", "Only side with door"),
    ("RoutingInformation (2 doors)", "{series} {hinge1} {swing1} / {series} {hinge2} {swing2}", "3082G.67P LH IS / 3082G.67P RH OS", "Both sides separated by /"),
    ("ProgramName (1 door)", "{series} {RH|LH} {O/S|I/S} {HINGE|KEEP}", "3082G.67P LH I/S KEEP", "HINGE if door hinges here, KEEP otherwise"),
    ("ProgramName (2 doors)", "{series} {seg1}, {seg2}", "3082G.67P LH I/S KEEP, RH O/S HINGE", "Comma-separated per side"),
    ("Hinge code", "RH if hinging_right=1, else LH", "", ""),
    ("Swing code", "O/S (or OS) if swinging_out=1, else I/S (or IS)", "", "RoutingInfo uses OS/IS; ProgramName uses O/S, I/S"),
    ("Attach code (ProgramName only)", "LEFT: HINGE if LD_hinging_right=1, else KEEP", "", "LEFT side: hinging_right=1 means door hinges here"),
    ("", "RIGHT: HINGE if RD_hinging_right=0, else KEEP", "", "RIGHT side: hinging_right=0 means door hinges here (inverted)"),
]

for field, pattern, example, notes in routing_ref:
    r += 1
    ws.cell(row=r, column=1, value=field).font = bold_font if field else Font()
    ws.cell(row=r, column=2, value=pattern)
    ws.cell(row=r, column=3, value=example)
    ws.cell(row=r, column=4, value=notes)
    style_row(ws, r, note_fill)

# ============================================================================
# SECTION 5: FLIP/ROTATE LOGIC
# ============================================================================
r += 2
ws.cell(row=r, column=1, value="FLIP / ROTATE LOGIC")
style_section(ws, r)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

r += 1
ws.cell(row=r, column=1, value="Condition").font = bold_font
ws.cell(row=r, column=2, value="Flip Value").font = bold_font
ws.cell(row=r, column=3, value="Meaning").font = bold_font
ws.cell(row=r, column=4, value="Example").font = bold_font
style_row(ws, r, note_fill)

flip_ref = [
    ("Stile: same-side drilling (both interior or both exterior)", "1", "Rotate", "left_int=1, right_int=1"),
    ("Stile: opposite-side drilling (left_int + right_ext OR left_ext + right_int)", "2", "Flip", "left_int=1, right_ext=1"),
    ("Door: always", "1", "Rotate", "All doors use Rotate"),
]

for cond, val, meaning, ex in flip_ref:
    r += 1
    ws.cell(row=r, column=1, value=cond)
    ws.cell(row=r, column=2, value=val)
    ws.cell(row=r, column=3, value=meaning)
    ws.cell(row=r, column=4, value=ex)
    style_row(ws, r, note_fill)

# ============================================================================
# SECTION 6: COLOR LEGEND
# ============================================================================
r += 2
ws.cell(row=r, column=1, value="COLOR LEGEND")
style_section(ws, r)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

legend = [
    (input_fill, "Yellow — Editable INPUT parameters (change these to test)"),
    (output_fill, "Blue — Direct mapping (input → output)"),
    (formula_fill, "Green — Derived/conditional formula (computed from inputs)"),
    (note_fill, "Gray — Constant or reference information"),
]

for fill, desc in legend:
    r += 1
    ws.cell(row=r, column=1, value="████")
    ws.cell(row=r, column=1).fill = fill
    ws.cell(row=r, column=2, value=desc)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)

# ── Freeze panes ────────────────────────────────────────────────────────────
ws.freeze_panes = "A5"

# ============================================================================
# ============================================================================
#  DOOR TAB
# ============================================================================
# ============================================================================
wd = wb.create_sheet("Door Cutlist Mapping")

# ── Column widths ───────────────────────────────────────────────────────────
wd.column_dimensions['A'].width = 40
wd.column_dimensions['B'].width = 22
wd.column_dimensions['C'].width = 18
wd.column_dimensions['D'].width = 45
wd.column_dimensions['E'].width = 55

# ============================================================================
# TITLE
# ============================================================================
r = 1
wd.cell(row=r, column=1, value="Voorwood Testing Planning (DOOR)")
wd.cell(row=r, column=1).font = Font(bold=True, size=14, color="2F5496")
wd.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

r = 2
wd.cell(row=r, column=1, value="Change yellow INPUT cells to see how CSV outputs update via formulas")
wd.cell(row=r, column=1).font = Font(italic=True, size=10, color="666666")
wd.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

# ============================================================================
# SECTION 1: INPUT PARAMETERS
# ============================================================================
r = 4
wd.cell(row=r, column=1, value="F360 MODEL INPUT PARAMETERS")
wd.cell(row=r, column=2, value="VALUE")
wd.cell(row=r, column=3, value="DATA TYPE")
wd.cell(row=r, column=4, value="DESCRIPTION")
wd.cell(row=r, column=5, value="EXAMPLE / NOTES")
style_header(wd, r)

door_inputs = [
    ("component_height", 81.1875, "float (in)", "Height of the door component", "Maps to Length in CSV"),
    ("component_width", 29.25, "float (in)", "Width of the door component", "Maps to Width in CSV"),
    ("component_thickness", 1.0, "float (in)", "Thickness of the door component", "Maps to Thickness in CSV"),
    ("ID", "1-D1-IBUS12345", "string", "Full component ID: {qty}-{comp_id}-{SO}", "Format: qty-compID-orderSO"),
    ("Series", "3082G.67P", "string", "Series identifier", "e.g. 3082G.67P, 3086G, etc."),
    ("component_floor_clearance", 1.0, "float (in)", "Floor clearance for the door", "Maps to FC (single value)"),
    ("door_hinging_right", 1, "bool (0/1)", "Door hinges on right side (viewed from outside)", "1=RH, 0=LH"),
    ("door_swinging_out", 1, "bool (0/1)", "Door swings outward into bathroom", "1=O/S, 0=I/S"),
    ("left_interior_rabbeting", 0, "bool (0/1)", "Left side has interior (bottom face) rabbeting", "From F360 model param"),
    ("left_exterior_rabbeting", 0, "bool (0/1)", "Left side has exterior (top face) rabbeting", "From F360 model param"),
    ("right_interior_rabbeting", 0, "bool (0/1)", "Right side has interior (bottom face) rabbeting", "From F360 model param"),
    ("right_exterior_rabbeting", 0, "bool (0/1)", "Right side has exterior (top face) rabbeting", "From F360 model param"),
]

d_input_start = 5
for i, (label, val, dtype, desc, notes) in enumerate(door_inputs):
    r = d_input_start + i
    wd.cell(row=r, column=1, value=label).font = bold_font
    wd.cell(row=r, column=2, value=val)
    wd.cell(row=r, column=3, value=dtype)
    wd.cell(row=r, column=4, value=desc)
    wd.cell(row=r, column=5, value=notes)
    style_row(wd, r, input_fill)
    wd.cell(row=r, column=2).font = Font(bold=True, size=11)

d_input_end = d_input_start + len(door_inputs) - 1

# Build cell ref dict
d_cells = {}
for i, (label, _, _, _, _) in enumerate(door_inputs):
    d_cells[label] = f"B{d_input_start + i}"

dc_height = d_cells["component_height"]
dc_width = d_cells["component_width"]
dc_thickness = d_cells["component_thickness"]
dc_id = d_cells["ID"]
dc_series = d_cells["Series"]
dc_fc = d_cells["component_floor_clearance"]
dc_hinging = d_cells["door_hinging_right"]
dc_swinging = d_cells["door_swinging_out"]
dc_left_int = d_cells["left_interior_rabbeting"]
dc_left_ext = d_cells["left_exterior_rabbeting"]
dc_right_int = d_cells["right_interior_rabbeting"]
dc_right_ext = d_cells["right_exterior_rabbeting"]

# ============================================================================
# SECTION 2: CSV OUTPUT VARIABLES
# ============================================================================
r = d_input_end + 2
d_output_section_start = r
wd.cell(row=r, column=1, value="VOORWOOD CSV OUTPUT (gCutlist)")
wd.cell(row=r, column=2, value="FORMULA VALUE")
wd.cell(row=r, column=3, value="MAPPING TYPE")
wd.cell(row=r, column=4, value="FORMULA / LOGIC")
wd.cell(row=r, column=5, value="CODE REFERENCE (parameter_exporter.py)")
style_header(wd, r)

door_outputs = [
    (
        "gCutlist.Item[0].Qty",
        "1",
        "Constant",
        "Always 1",
        "Line 1033: ('Qty', '1')"
    ),
    (
        "gCutlist.Item[0].Width",
        f"={dc_width}",
        "Direct",
        "component_width",
        "Line 902: component_width"
    ),
    (
        "gCutlist.Item[0].Length",
        f"={dc_height}",
        "Direct",
        "component_height (door height = CSV Length)",
        "Line 905: component_height"
    ),
    (
        "gCutlist.Item[0].ID",
        f'=MID({dc_id},FIND("-",{dc_id})+1,FIND("-",{dc_id},FIND("-",{dc_id})+1)-FIND("-",{dc_id})-1)',
        "Extracted",
        'Extract component ID from full ID (e.g. "D1" from "1-D1-IBUS12345")',
        "Line 1036: component_id"
    ),
    (
        "gCutlist.Item[0].Series",
        f"={dc_series}",
        "Direct",
        "Series ID passed through",
        "Line 1037: series_id"
    ),
    (
        "gCutlist.Item[0].RoutingInformation",
        f'={dc_series}&" "&IF({dc_hinging}=1,"RH","LH")&" "&IF({dc_swinging}=1,"OS","IS")',
        "Derived",
        '{series} {RH|LH} {OS|IS} — single door, no per-side logic',
        "Lines 956-961: door routing = series + hinge + swing"
    ),
    (
        "gCutlist.Item[0].SO",
        f'=MID({dc_id},FIND("-",{dc_id},FIND("-",{dc_id})+1)+1,LEN({dc_id}))',
        "Extracted",
        'Extract order/SO from full ID (e.g. "IBUS12345" from "1-D1-IBUS12345")',
        "Line 1039: order_id"
    ),
    (
        "gCutlist.Item[0].FC",
        f'={dc_fc}',
        "Direct",
        "component_floor_clearance (single value for doors, not per-side)",
        "Lines 931-936: component_floor_clearance from JSON or model"
    ),
    (
        "gCutlist.Item[0].GappingSide1",
        f'=IF(OR({dc_right_int}=1,{dc_right_ext}=1),1,0)',
        "Derived",
        "1 if right side has any rabbeting (interior OR exterior), else 0",
        "Lines 970-972: has_right_rabbeting"
    ),
    (
        "gCutlist.Item[0].GappingSide2",
        f'=IF(OR({dc_left_int}=1,{dc_left_ext}=1),1,0)',
        "Derived",
        "1 if left side has any rabbeting (interior OR exterior), else 0",
        "Lines 970-973: has_left_rabbeting"
    ),
    (
        "gCutlist.Item[0].GappingSide3",
        "0",
        "Constant",
        "Always 0 (top edge — no rabbeting)",
        "Line 974: gapping_side3 = 0"
    ),
    (
        "gCutlist.Item[0].GappingSide4",
        "0",
        "Constant",
        "Always 0 (bottom edge — no rabbeting)",
        "Line 975: gapping_side4 = 0"
    ),
    (
        "gCutlist.Item[0].GappingStartSide1",
        "0",
        "Constant",
        "Door: always 0 (rabbeting runs full length, no offset)",
        "Line 994: gapping_start_side1 = 0"
    ),
    (
        "gCutlist.Item[0].GappingStartSide2",
        "0",
        "Constant",
        "Door: always 0 (rabbeting runs full length, no offset)",
        "Line 996: gapping_start_side2 = 0"
    ),
    (
        "gCutlist.Item[0].GappingLengthSide1",
        f'=IF(OR({dc_right_int}=1,{dc_right_ext}=1),{dc_height},0)',
        "Conditional",
        "component_height if right side has rabbeting, else 0 (full-length rabbeting)",
        "Line 995: length_value if gapping_side1"
    ),
    (
        "gCutlist.Item[0].GappingLengthSide2",
        f'=IF(OR({dc_left_int}=1,{dc_left_ext}=1),{dc_height},0)',
        "Conditional",
        "component_height if left side has rabbeting, else 0 (full-length rabbeting)",
        "Line 997: length_value if gapping_side2"
    ),
    (
        "gCutlist.Item[0].GappingStartSide3",
        "0",
        "Constant",
        "N/A (GappingSide3 is always 0)",
        "Line 998"
    ),
    (
        "gCutlist.Item[0].GappingStartSide4",
        "0",
        "Constant",
        "N/A (GappingSide4 is always 0)",
        "Line 999"
    ),
    (
        "gCutlist.Item[0].GappingLengthSide3",
        "0",
        "Constant",
        "N/A (GappingSide3 is always 0)",
        "Line 1000"
    ),
    (
        "gCutlist.Item[0].GappingLengthSide4",
        "0",
        "Constant",
        "N/A (GappingSide4 is always 0)",
        "Line 1001"
    ),
    (
        "gCutlist.Item[0].GappingTopSide1",
        f'=IF(OR({dc_right_int}=1,{dc_right_ext}=1),IF({dc_right_ext}=1,1,0),0)',
        "Conditional",
        "1 if right side has exterior rabbeting, 0 if interior only, 0 if no rabbeting",
        "Lines 1017-1020: right_exterior_rabbeting if gapping_side1"
    ),
    (
        "gCutlist.Item[0].GappingTopSide2",
        f'=IF(OR({dc_left_int}=1,{dc_left_ext}=1),IF({dc_left_ext}=1,1,0),0)',
        "Conditional",
        "1 if left side has exterior rabbeting, 0 if interior only, 0 if no rabbeting",
        "Lines 1021-1024: left_exterior_rabbeting if gapping_side2"
    ),
    (
        "gCutlist.Item[0].GappingTopSide3",
        "0",
        "Constant",
        "N/A (GappingSide3 is always 0)",
        "Line 1025"
    ),
    (
        "gCutlist.Item[0].GappingTopSide4",
        "0",
        "Constant",
        "N/A (GappingSide4 is always 0)",
        "Line 1026"
    ),
    (
        "gCutlist.Item[0].EdgeFinishingSide1",
        f'=IF(OR({dc_right_int}=1,{dc_right_ext}=1),0,1)',
        "Derived",
        "Inverse of GappingSide1: 0 if rabbeting on right, else 1",
        "Line 978: edge_finishing_side1 = 0 if gapping_side1 else 1"
    ),
    (
        "gCutlist.Item[0].EdgeFinishingSide2",
        f'=IF(OR({dc_left_int}=1,{dc_left_ext}=1),0,1)',
        "Derived",
        "Inverse of GappingSide2: 0 if rabbeting on left, else 1",
        "Line 979: edge_finishing_side2 = 0 if gapping_side2 else 1"
    ),
    (
        "gCutlist.Item[0].EdgeFinishingSide3",
        "1",
        "Constant",
        "Always 1 (top edge always edge-finished)",
        "Line 980: edge_finishing_side3 = 1"
    ),
    (
        "gCutlist.Item[0].EdgeFinishingSide4",
        "1",
        "Constant",
        "Always 1 (bottom edge always edge-finished)",
        "Line 981: edge_finishing_side4 = 1"
    ),
    (
        "gCutlist.Item[0].Thickness",
        f"={dc_thickness}",
        "Direct",
        "component_thickness",
        "Line 1029: component_thickness"
    ),
    (
        "gCutlist.Item[0].Flip",
        "1",
        "Constant",
        "Door: always 1 (Rotate). Doors never use Flip.",
        "Line 1062: ('Flip', '1') — always 1 for doors"
    ),
    (
        "gCutlist.Item[0].Location",
        "BLA",
        "Constant",
        "Always BLA",
        "Line 1063: ('Location', 'BLA')"
    ),
    (
        "gCutlist.Item[0].ProgramName",
        f'={dc_series}&" "&IF({dc_hinging}=1,"RH","LH")&" "&IF({dc_swinging}=1,"O/S","I/S")&" DOOR"',
        "Derived",
        '{series} {RH|LH} {O/S|I/S} DOOR',
        "Lines 719-725: drill_name = series + hinge + swing + DOOR"
    ),
]

# Write door outputs
d_output_start = d_output_section_start + 1
for i, (csv_var, formula, map_type, logic, code_ref) in enumerate(door_outputs):
    r = d_output_start + i
    wd.cell(row=r, column=1, value=csv_var).font = bold_font
    wd.cell(row=r, column=2, value=formula if not formula.startswith("=") else None)
    if formula.startswith("="):
        wd.cell(row=r, column=2).value = formula
    wd.cell(row=r, column=3, value=map_type)
    wd.cell(row=r, column=4, value=logic)
    wd.cell(row=r, column=5, value=code_ref)

    if map_type == "Constant":
        style_row(wd, r, note_fill)
    elif map_type == "Direct":
        style_row(wd, r, output_fill)
    elif map_type in ("Derived", "Conditional", "Extracted"):
        style_row(wd, r, formula_fill)
    else:
        style_row(wd, r, output_fill)

d_output_end = d_output_start + len(door_outputs) - 1

# ============================================================================
# SECTION 3: KEY DIFFERENCES FROM STILE
# ============================================================================
r = d_output_end + 2
wd.cell(row=r, column=1, value="KEY DIFFERENCES FROM STILE")
style_section(wd, r)
wd.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

r += 1
wd.cell(row=r, column=1, value="Field").font = bold_font
wd.cell(row=r, column=2, value="Stile").font = bold_font
wd.cell(row=r, column=3, value="Door").font = bold_font
wd.cell(row=r, column=4, value="Notes").font = bold_font
style_row(wd, r, note_fill)

diff_data = [
    ("FC", "Per-side: LD_floor_clearance / RD_floor_clearance", "Single: component_floor_clearance", "Stile has FC SIDE 1 and FC SIDE 2; Door has one FC"),
    ("RoutingInformation", "{series} {RH|LH} {OS|IS} per side, joined /", "{series} {RH|LH} {OS|IS}", "Door has single routing, no per-side"),
    ("ProgramName", "{series} {hinge} {swing} {HINGE|KEEP} per side", "{series} {RH|LH} {O/S|I/S} DOOR", "Door always ends with DOOR"),
    ("GappingSide1/2", "1 if that side has a door", "1 if that side has any rabbeting (int OR ext)", "Different trigger condition"),
    ("GappingStart", "right/left_rabbeting_top from model", "Always 0", "Door rabbeting has no offset"),
    ("GappingLength", "right/left_rabbeting_length from model", "component_height if gapping, else 0", "Door rabbeting runs full length"),
    ("Flip", "1 or 2 (based on opposite drilling)", "Always 1 (Rotate)", "Doors never flip"),
    ("Input params", "left_side_door, LD_hinging_right, LD_swinging_out, ...", "door_hinging_right, door_swinging_out", "Different JSON param names"),
]

for field, stile_val, door_val, notes in diff_data:
    r += 1
    wd.cell(row=r, column=1, value=field).font = bold_font
    wd.cell(row=r, column=2, value=stile_val)
    wd.cell(row=r, column=3, value=door_val)
    wd.cell(row=r, column=4, value=notes)
    style_row(wd, r, note_fill)

# ============================================================================
# SECTION 4: COLOR LEGEND
# ============================================================================
r += 2
wd.cell(row=r, column=1, value="COLOR LEGEND")
style_section(wd, r)
wd.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

for fill, desc in legend:
    r += 1
    wd.cell(row=r, column=1, value="████")
    wd.cell(row=r, column=1).fill = fill
    wd.cell(row=r, column=2, value=desc)
    wd.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)

# ── Freeze panes ────────────────────────────────────────────────────────────
wd.freeze_panes = "A5"

# ── Save ────────────────────────────────────────────────────────────────────
output_path = r"c:\Users\james.derrod\VM Fusion Extension\testing_suite\Voorwood_Stile_Cutlist_Mapping.xlsx"
wb.save(output_path)
print(f"Spreadsheet created: {output_path}")
